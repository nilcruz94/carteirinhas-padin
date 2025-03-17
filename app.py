from flask import Flask, request, redirect, url_for, render_template_string, jsonify, session, flash, send_file
import pandas as pd
import os
import uuid
import re
from datetime import datetime
from io import BytesIO
from werkzeug.utils import secure_filename
from functools import wraps
import locale
import xlrd  # Para ler arquivos .xls, se necessário
from openpyxl import load_workbook, Workbook  # Usado para trabalhar com XLSX
from openpyxl.utils import get_column_letter  # Para obter a coluna em letra
from openpyxl.cell import MergedCell  # Para identificar células mescladas
from flask import Flask, request, redirect, url_for, render_template, render_template_string, jsonify, session, flash, send_file

# Tenta definir a localidade para formatação de datas em português
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    pass

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta'  # Altere para uma chave segura
ACCESS_TOKEN = "minha_senha"  # Token de acesso

app.config['UPLOAD_FOLDER'] = 'uploads'
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif'}

# Cria os diretórios necessários, se não existirem
if not os.path.exists('static/fotos'):
    os.makedirs('static/fotos')
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])


def allowed_file(filename):
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated_function


# Função para atualizar valor em célula mesclada (mantém a mesclagem)
def set_merged_cell_value(ws, cell_coord, value):
    cell = ws[cell_coord]
    if isinstance(cell, MergedCell):
        # Procura o intervalo mesclado que contém a célula
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                # Obtém a célula superior esquerda do intervalo mesclado
                min_col, min_row, _, _ = merged_range.bounds
                top_left_coord = f"{get_column_letter(min_col)}{min_row}"
                ws[top_left_coord] = value
                ws.merge_cells(range_str)
                return
    ws[cell_coord] = value


def convert_xls_to_xlsx(file_like):
    """
    Converte um arquivo XLS (file-like) para um Workbook do openpyxl.
    """
    book_xlrd = xlrd.open_workbook(file_contents=file_like.read())
    wb = Workbook()
    # Remover a planilha padrão criada pelo openpyxl, se houver
    if "Sheet" in wb.sheetnames and len(book_xlrd.sheet_names()) > 0:
        std = wb.active
        wb.remove(std)
    for sheet_name in book_xlrd.sheet_names():
        sheet_xlrd = book_xlrd.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row in range(sheet_xlrd.nrows):
            for col in range(sheet_xlrd.ncols):
                ws.cell(row=row+1, column=col+1, value=sheet_xlrd.cell_value(row, col))
    return wb


def load_workbook_model(file):
    """
    Abre o arquivo do modelo XLSX (ou XLS convertendo-o para XLSX) preservando toda a formatação.
    """
    ext = os.path.splitext(file.filename)[1].lower()
    file.seek(0)
    if ext == '.xlsx':
        return load_workbook(file, data_only=False)
    elif ext == '.xls':
        content = file.read()
        return convert_xls_to_xlsx(BytesIO(content))
    else:
        raise ValueError("Formato de arquivo não suportado para o quadro modelo.")


def gerar_html_carteirinhas(arquivo_excel):
    planilha = pd.read_excel(arquivo_excel, sheet_name='LISTA CORRIDA')
    dados = planilha[['RM', 'NOME', 'DATA NASC.', 'RA', 'SAI SOZINHO?', 'SÉRIE', 'HORÁRIO']]
    dados['RM'] = dados['RM'].fillna(0).astype(int)

    alunos_sem_fotos_list = []
    html_content = """
<!DOCTYPE html>
<html lang="pt-br">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Carteirinhas - E.M José Padin Mouta</title>
  <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.4/css/all.min.css">
  <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
  <style>
    /* Estilos CSS para carteirinhas */
    body {
      font-family: 'Montserrat', sans-serif;
      margin: 0;
      padding: 20px;
      background-color: #f4f4f4;
      display: flex;
      flex-direction: column;
      align-items: center;
    }
    /* Demais estilos omitidos para brevidade */
  </style>
</head>
<body>
  <!-- Conteúdo HTML gerado dinamicamente -->
</body>
</html>
"""
    return render_template_string(html_content)


def gerar_declaracao_escolar(file_path, rm, tipo, file_path2=None):
    if session.get('declaracao_tipo') != "EJA":
        planilha = pd.read_excel(file_path, sheet_name='LISTA CORRIDA')

        def format_rm(x):
            try:
                return str(int(float(x)))
            except:
                return str(x)

        planilha['RM_str'] = planilha['RM'].apply(format_rm)
        try:
            rm_num = str(int(float(rm)))
        except:
            rm_num = str(rm)
        aluno = planilha[planilha['RM_str'] == rm_num]
        if aluno.empty:
            return None
        row = aluno.iloc[0]
        nome = row['NOME']
        serie = row['SÉRIE']
        if isinstance(serie, str):
            serie = re.sub(r"(\d+º)([A-Za-z])", r"\1 ano \2", serie)
        data_nasc = row['DATA NASC.']
        ra = row['RA']
        horario = row['HORÁRIO']
        if pd.isna(horario) or not str(horario).strip():
            horario = "Desconhecido"
        else:
            horario = str(horario).strip()
        ra_label = "RA"
        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors='coerce')
                if pd.notna(data_nasc):
                    data_nasc = data_nasc.strftime('%d/%m/%Y')
                else:
                    data_nasc = "Desconhecida"
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"
    else:
        df = pd.read_excel(file_path, sheet_name=0, header=None, skiprows=1)
        df['RM_str'] = df.iloc[:, 2].apply(lambda x: str(int(x)) if pd.notna(x) and float(x) != 0 else "")
        df['NOME'] = df.iloc[:, 3]
        df['NASC.'] = df.iloc[:, 6]
        def get_ra(row):
            try:
                val = row.iloc[7]
                if pd.isna(val) or float(val) == 0:
                    return row.iloc[8]
                else:
                    return val
            except:
                return row.iloc[7]
        df['RA'] = df.apply(get_ra, axis=1)
        df['SÉRIE'] = df.iloc[:, 0]
        try:
            rm_num = str(int(float(rm)))
        except:
            rm_num = str(rm)
        aluno = df[df['RM_str'] == rm_num]
        if aluno.empty:
            return None
        row = aluno.iloc[0]
        nome = row['NOME']
        serie = row['SÉRIE']
        if isinstance(serie, str):
            serie = re.sub(r"(\d+º)([A-Za-z])", r"\1 ano \2", serie)
        data_nasc = row['NASC.']
        ra = row['RA']
        original_ra = row.iloc[7]
        if pd.isna(original_ra) or (isinstance(original_ra, (int, float)) and float(original_ra) == 0):
            ra_label = "RG"
        else:
            ra_label = "RA"
        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors='coerce')
                if pd.notna(data_nasc):
                    data_nasc = data_nasc.strftime('%d/%m/%Y')
                else:
                    data_nasc = "Desconhecida"
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"

    now = datetime.now()
    meses = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio", 6: "junho",
             7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
    mes = meses[now.month].capitalize()
    data_extenso = f"Praia Grande, {now.day:02d} de {mes} de {now.year}"

    additional_css = '''
.print-button {
  background-color: #283E51;
  color: #fff;
  border: none;
  padding: 10px 20px;
  border-radius: 5px;
  cursor: pointer;
  margin-top: 20px;
}
.print-button:hover {
  background-color: #1d2d3a;
}
'''

    if tipo == "Escolaridade":
        titulo = "Declaração de Escolaridade"
        if session.get('declaracao_tipo') == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do {ra_label} <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"encontra-se regularmente matriculado(a) na E.M José Padin Mouta, cursando atualmente o "
                f"<strong><u>{serie}</u></strong>."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do RA <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"encontra-se regularmente matriculado(a) na E.M José Padin Mouta, cursando atualmente o "
                f"<strong><u>{serie}</u></strong> no horário de aula: <strong><u>{horario}</u></strong>."
            )

    elif tipo == "Transferencia":
        titulo = "Declaração de Transferência"
        if session.get('declaracao_tipo') == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do {ra_label} <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"solicitou transferência de nossa unidade escolar na data de hoje, estando apto(a) a cursar o "
                f"<strong><u>{serie}</u></strong>."
            )
        else:
            serie_mod = re.sub(r"^(\d+º).*", r"\1 ano", serie)
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) responsável do(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do RA <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"compareceu a nossa unidade escolar e solicitou transferência na data de hoje, o aluno está apto(a) a cursar o "
                f"<strong><u>{serie_mod}</u></strong>."
            )

    elif tipo == "Conclusão":
        titulo = "Declaração de Conclusão"
        if session.get('declaracao_tipo') == "EJA":
            mapping = {
                "1ª SÉRIE E.F": "2ª SÉRIE E.F",
                "2ª SÉRIE E.F": "3ª SÉRIE E.F",
                "3ª SÉRIE E.F": "4ª SÉRIE E.F",
                "4ª SÉRIE E.F": "5ª SÉRIE E.F",
                "5ª SÉRIE E.F": "6ª SÉRIE E.F",
                "6ª SÉRIE E.F": "7ª SÉRIE E.F",
                "7ª SÉRIE E.F": "8ª SÉRIE E.F",
                "8ª SÉRIE E.F": "1ª SÉRIE E.M",
                "1ª SÉRIE E.M": "2ª SÉRIE E.M",
                "2ª SÉRIE E.M": "3ª SÉRIE E.M",
                "3ª SÉRIE E.M": "ENSINO SUPERIOR"
            }
            series_text = mapping.get(serie, "a série subsequente")
        else:
            match = re.search(r"(\d+)º\s*ano", serie)
            if match:
                next_year = int(match.group(1)) + 1
                series_text = f"{next_year}º ano"
            else:
                series_text = "a série subsequente"

        if session.get('declaracao_tipo') == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do {ra_label} <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"concluiu com êxito o <strong><u>{serie}</u></strong>, estando apto(a) a ingressar no "
                f"<strong><u>{series_text}</u></strong>."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) <strong><u>{nome}</u></strong>, "
                f"portador(a) do RA <strong><u>{ra}</u></strong>, nascido(a) em <strong><u>{data_nasc}</u></strong>, "
                f"concluiu com êxito o <strong><u>{serie}</u></strong>, estando apto(a) a ingressar no "
                f"<strong><u>{series_text}</u></strong>."
            )

    else:
        titulo = "Declaração"
        declaracao_text = "Tipo de declaração inválido."

    base_template = f'''<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <title>{titulo} - E.M José Padin Mouta</title>
  <style>
    @page {{
      margin: 0;
    }}
    html, body {{
      margin: 0;
      padding: 0.5cm;
      font-family: 'Montserrat', sans-serif;
      font-size: 16px;
      line-height: 1.5;
      color: #333;
    }}
    .header {{
      text-align: center;
      border-bottom: 2px solid #283E51;
      padding-bottom: 5px;
      margin-bottom: 10px;
    }}
    .header h1 {{
      margin: 0;
      font-size: 24px;
      text-transform: uppercase;
      color: #283E51;
    }}
    .header p {{
      margin: 3px 0;
      font-size: 16px;
    }}
    .date {{
      text-align: right;
      font-size: 16px;
      margin-bottom: 10px;
    }}
    .content {{
      text-align: justify;
      margin-bottom: 10px;
    }}
    .signature {{
      text-align: center;
      margin: 0;
      padding: 0;
    }}
    .signature .line {{
      height: 1px;
      background-color: #333;
      width: 60%;
      margin: 0 auto 5px auto;
    }}
    .footer {{
      text-align: center;
      border-top: 2px solid #283E51;
      padding-top: 5px;
      margin: 0;
      font-size: 14px;
      color: #555;
    }}
    @media print {{
      .no-print {{ display: none !important; }}
      body {{
        margin: 0;
        padding: 0.5cm;
        font-size: 16px;
      }}
      .declaration-bottom {{
         margin-top: 10cm;
      }}
      .date {{
         margin-top: 2cm;
      }}
    }}
    {additional_css}
    header {{
      background: linear-gradient(90deg, #283E51, #4B79A1);
      color: #fff;
      padding: 20px;
      text-align: center;
      border-bottom: 3px solid #1d2d3a;
      border-radius: 0 0 15px 15px;
      box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }}
  </style>
</head>
<body>
  <div class="declaration-container">
    <div class="header">
      <div style="display: flex; justify-content: space-between; align-items: center;">
        <img src="/static/logos/escola.png" alt="Escola Logo" style="height: 80px;">
        <div>
          <h1>Secretaria de Educação</h1>
          <p>E.M José Padin Mouta</p>
          <p>Município da Estância Balneária de Praia Grande</p>
          <p>Estado de São Paulo</p>
        </div>
        <img src="/static/logos/municipio.png" alt="Município Logo" style="height: 80px;">
      </div>
    </div>
    <div class="date">
      <p>{data_extenso}</p>
    </div>
    <div class="content">
      <h2 style="text-align: center; text-transform: uppercase; color: #283E51;">{titulo}</h2>
      <p>{declaracao_text}</p>
    </div>
    <div class="declaration-bottom">
      <div class="signature">
        <div class="line"></div>
        <p>Luciana Rocha Augustinho</p>
        <p>Diretora da Unidade Escolar</p>
      </div>
      <div class="footer">
        <p>Rua: Bororós, nº 150, Vila Tupi, Praia Grande - SP, CEP: 11703-390</p>
        <p>Telefone: 3496-5321 | E-mail: em.padin@praiagrande.sp.gov.br</p>
      </div>
    </div>
  </div>
  <div class="no-print" style="text-align: center; margin-top: 20px;">
    <button onclick="window.print()" class="print-button">Imprimir Declaração</button>
  </div>
</body>
</html>
'''
    return base_template


@app.route('/login', methods=['GET', 'POST'])
def login_route():
    error = None
    if request.method == 'POST':
        token = request.form.get('token')
        if token == ACCESS_TOKEN:
            session['logged_in'] = True
            if 'lista_fundamental' not in session or 'lista_eja' not in session:
                return redirect(url_for('upload_listas'))
            return redirect(url_for('dashboard'))
        else:
            error = "Token inválido. Tente novamente."
    login_html = '''
    <!doctype html>
    <html lang="pt-br">
      <head>
        <meta charset="utf-8">
        <title>Login - Acesso Restrito</title>
        <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <style>
          body {
            background: linear-gradient(135deg, #283E51, #4B79A1);
            font-family: 'Montserrat', sans-serif;
            margin: 0;
            padding: 0;
            display: flex;
            flex-direction: column;
            min-height: 100vh;
          }
          header {
            background: linear-gradient(90deg, #283E51, #4B79A1);
            color: #fff;
            padding: 20px;
            text-align: center;
            border-bottom: 3px solid #1d2d3a;
            border-radius: 0 0 15px 15px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
          }
          main {
            flex: 1;
            display: flex;
            align-items: center;
            justify-content: center;
          }
          .container-login {
            background: #fff;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            width: 100%;
            max-width: 400px;
          }
          .container-login h2 {
            margin-bottom: 20px;
            font-weight: 600;
            color: #283E51;
          }
          .btn-primary {
            background-color: #283E51;
            border: none;
          }
          .btn-primary:hover {
            background-color: #1d2d3a;
          }
          footer {
            background-color: #424242;
            color: #fff;
            text-align: center;
            padding: 10px;
          }
          .error {
            color: red;
            margin-top: 15px;
          }
        </style>
      </head>
      <body>
        <header>
          <h1 class="mb-0">Secretaria - E.M José Padin Mouta</h1>
        </header>
        <main>
          <div class="container container-login">
            <h2>Acesso Restrito</h2>
            <form method="POST">
              <div class="form-group">
                <input type="password" name="token" class="form-control" placeholder="Digite o token de acesso" required>
              </div>
              <button type="submit" class="btn btn-primary btn-block">Entrar</button>
            </form>
            {% if error %}
              <p class="error">{{ error }}</p>
            {% endif %}
          </div>
        </main>
        <footer>
          Desenvolvido por Nilson Cruz © 2025. Todos os direitos reservados.
        </footer>
      </body>
    </html>
    '''
    return render_template_string(login_html, error=error)


@app.route('/logout')
def logout_route():
    session.clear()
    return redirect(url_for('login_route'))


@app.route('/upload_listas', methods=['GET', 'POST'])
@login_required
def upload_listas():
    if request.method == 'POST':
        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')
        if not fundamental_file or fundamental_file.filename == '':
            flash("Selecione a Lista Piloto - REGULAR - 2025", "error")
            return redirect(url_for('upload_listas'))
        if not eja_file or eja_file.filename == '':
            flash("Selecione a Lista Piloto - EJA - 1º SEM - 2025", "error")
            return redirect(url_for('upload_listas'))
        fundamental_filename = secure_filename(f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename)
        eja_filename = secure_filename(f"eja_{uuid.uuid4().hex}_" + eja_file.filename)
        fundamental_path = os.path.join(app.config['UPLOAD_FOLDER'], fundamental_filename)
        eja_path = os.path.join(app.config['UPLOAD_FOLDER'], eja_filename)
        fundamental_file.save(fundamental_path)
        eja_file.save(eja_path)
        session['lista_fundamental'] = fundamental_path
        session['lista_eja'] = eja_path
        flash("Listas carregadas com sucesso.", "success")
        return redirect(url_for('dashboard'))
    upload_listas_html = '''
    <!doctype html>
    <html lang="pt-br">
    <head>
      <meta charset="utf-8">
      <title>Upload de Listas Piloto</title>
      <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
      <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
      <style>
      body {
          background: #eef2f3;
          font-family: 'Montserrat', sans-serif;
      }
      header {
          background: linear-gradient(90deg, #283E51, #4B79A1);
          color: #fff;
          padding: 20px;
          text-align: center;
          border-bottom: 3px solid #1d2d3a;
          border-radius: 0 0 15px 15px;
          box-shadow: 0 4px 6px rgba(0,0,0,0.1);
      }
      .container-form {
          background: #fff;
          padding: 40px;
          border-radius: 10px;
          box-shadow: 0 4px 12px rgba(0,0,0,0.15);
          margin: 40px auto;
          max-width: 600px;
      }
      .btn-primary {
          background-color: #283E51;
          border: none;
      }
      .btn-primary:hover {
          background-color: #1d2d3a;
      }
      footer {
          background-color: #424242;
          color: #fff;
          text-align: center;
          padding: 10px;
          position: fixed;
          bottom: 0;
          width: 100%;
      }
      </style>
    </head>
    <body>
      <header>
        <h1>Upload de Listas Piloto</h1>
      </header>
      <div class="container-form">
        <form method="POST" enctype="multipart/form-data">
          <div class="form-group">
            <label for="lista_fundamental">Selecione a Lista Piloto - REGULAR - 2025:</label>
            <input type="file" class="form-control-file" name="lista_fundamental" id="lista_fundamental" accept=".xlsx, .xls" required>
          </div>
          <div class="form-group">
            <label for="lista_eja">Selecione a Lista Piloto - EJA - 1º SEM - 2025:</label>
            <input type="file" class="form-control-file" name="lista_eja" id="lista_eja" accept=".xlsx, .xls" required>
          </div>
          <button type="submit" class="btn btn-primary">Carregar Listas</button>
        </form>
      </div>
      <footer>
        Desenvolvido por Nilson Cruz © 2025. Todos os direitos reservados.
      </footer>
    </body>
    </html>
    '''
    return render_template_string(upload_listas_html)


@app.route('/', methods=['GET'])
@login_required
def dashboard():
    dashboard_html = '''
    <!doctype html>
    <html lang="pt-br">
      <head>
        <meta charset="utf-8">
        <title>Secretaria - E.M José Padin Mouta</title>
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.4/css/all.min.css">
        <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
        <style>
          body {
            background: #eef2f3;
            font-family: 'Montserrat', sans-serif;
          }
          header {
            background: linear-gradient(90deg, #283E51, #4B79A1);
            color: #fff;
            padding: 20px;
            text-align: center;
            border-bottom: 3px solid #1d2d3a;
            border-radius: 0 0 15px 15px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
          }
          .container-dashboard {
            background: #fff;
            padding: 40px;
            border-radius: 10px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
            margin: 40px auto;
            max-width: 800px;
          }
          .option-card {
            border: 1px solid #ccc;
            border-radius: 10px;
            padding: 20px;
            text-align: center;
            transition: transform 0.2s;
            cursor: pointer;
            margin-bottom: 20px;
          }
          .option-card:hover {
            transform: scale(1.02);
          }
          .option-card h2 {
            margin-bottom: 10px;
            color: #283E51;
          }
          .option-card p {
            color: #555;
          }
          .logout-container {
            text-align: center;
            margin-top: 20px;
          }
          .btn-logout {
            background-color: #dc3545;
            color: #fff;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            font-size: 16px;
            text-decoration: none;
            transition: background-color 0.3s;
          }
          .btn-logout:hover {
            background-color: #c82333;
          }
          footer {
            background-color: #424242;
            color: #fff;
            text-align: center;
            padding: 10px;
            position: fixed;
            bottom: 0;
            width: 100%;
          }
        </style>
      </head>
      <body>
        <header>
          <h1>Secretaria - E.M José Padin Mouta</h1>
        </header>
        <div class="container container-dashboard">
          <div class="option-card" onclick="window.location.href='{{ url_for('declaracao_tipo') }}'">
            <h2>Declaração Escolar</h2>
            <p>Gerar declaração escolar.</p>
          </div>
          <div class="option-card" onclick="window.location.href='{{ url_for('carteirinhas') }}'">
            <h2>Carteirinhas</h2>
            <p>Gerar carteirinhas para os alunos.</p>
          </div>
          <div class="option-card" onclick="window.location.href='{{ url_for('quadros') }}'">
            <h2>Quadros</h2>
            <p>Gerar quadros para os alunos.</p>
          </div>
          <div class="logout-container">
            <a href="{{ url_for('logout_route') }}" class="btn-logout">
              <i class="fas fa-sign-out-alt"></i> Logout
            </a>
          </div>
        </div>
        <footer>
          Desenvolvido por Nilson Cruz © 2025. Todos os direitos reservados.
        </footer>
      </body>
    </html>
    '''
    return render_template_string(dashboard_html)


@app.route('/carteirinhas', methods=['GET', 'POST'])
@login_required
def carteirinhas():
    if request.method == 'POST':
        file = None
        if 'excel_file' in request.files and request.files['excel_file'].filename != '':
            file = request.files['excel_file']
            filename = secure_filename(file.filename)
            unique_filename = f"carteirinhas_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            session['lista_fundamental'] = file_path
            file = open(file_path, 'rb')
        else:
            file_path = session.get('lista_fundamental')
            if file_path and os.path.exists(file_path):
                file = open(file_path, 'rb')
        if not file:
            return "Nenhum arquivo selecionado", 400
        flash("Gerando carteirinhas. Aguarde...", "info")
        html_result = gerar_html_carteirinhas(file)
        file.close()
        return html_result
    carteirinhas_html = '''
    <!doctype html>
    <html lang="pt-br">
      <head>
        <meta charset="utf-8">
        <title>Carteirinhas - E.M José Padin Mouta</title>
        <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.4/css/all.min.css">
        <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
        <style>
          /* Estilos para a página de carteirinhas (omitidos para brevidade) */
        </style>
      </head>
      <body>
        <!-- Conteúdo HTML da página de carteirinhas -->
      </body>
    </html>
    '''
    return render_template_string(carteirinhas_html)


@app.route('/declaracao/upload', methods=['GET', 'POST'])
@login_required
def declaracao_upload():
    if request.method == 'POST':
        file = None
        if 'excel_file' in request.files and request.files['excel_file'].filename != '':
            file = request.files['excel_file']
            filename = secure_filename(file.filename)
            unique_filename = f"declaracao_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            session['lista_fundamental'] = file_path
        else:
            file_path = session.get('lista_fundamental')
            if file_path and os.path.exists(file_path):
                file = open(file_path, 'rb')
        if not file:
            flash("Nenhum arquivo enviado.", "error")
            return redirect(url_for('declaracao_upload'))
        session['declaracao_excel'] = file_path
        session['declaracao_tipo'] = "Fundamental"
        if hasattr(file, 'close'):
            file.close()
        return redirect(url_for('declaracao_select'))
    upload_form = '''
    <!doctype html>
    <html lang="pt-br">
    <head>
      <meta charset="utf-8">
      <title>Declaração Escolar - Fundamental</title>
      <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
      <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
      <style>
        /* Estilos omitidos para brevidade */
      </style>
    </head>
    <body>
      <!-- HTML da página de upload para Fundamental -->
    </body>
    </html>
    '''
    return render_template_string(upload_form)


@app.route('/declaracao/upload_eja', methods=['GET', 'POST'])
@login_required
def declaracao_upload_eja():
    if request.method == 'POST':
        file = None
        if 'excel_file' in request.files and request.files['excel_file'].filename != '':
            file = request.files['excel_file']
            filename = secure_filename(file.filename)
            unique_filename = f"declaracao2_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            session['lista_eja'] = file_path
        else:
            file_path = session.get('lista_eja')
            if file_path and os.path.exists(file_path):
                file = open(file_path, 'rb')
        if not file:
            flash("Nenhum arquivo enviado.", "error")
            return redirect(url_for('declaracao_upload_eja'))
        session['declaracao_excel'] = file_path
        session['declaracao_tipo'] = "EJA"
        if hasattr(file, 'close'):
            file.close()
        return redirect(url_for('declaracao_select'))
    upload_form = '''
    <!doctype html>
    <html lang="pt-br">
    <head>
      <meta charset="utf-8">
      <title>Declaração Escolar - EJA</title>
      <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
      <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
      <style>
        /* Estilos omitidos para brevidade */
      </style>
    </head>
    <body>
      <!-- HTML da página de upload para EJA -->
    </body>
    </html>
    '''
    return render_template_string(upload_form)


@app.route('/declaracao/select', methods=['GET', 'POST'])
@login_required
def declaracao_select():
    file_path = session.get('declaracao_excel')
    if not file_path or not os.path.exists(file_path):
        flash("Arquivo Excel não encontrado. Por favor, anexe a lista piloto.", "error")
        if session.get('declaracao_tipo') == "EJA":
            return redirect(url_for('declaracao_upload_eja'))
        else:
            return redirect(url_for('declaracao_upload'))
    if session.get('declaracao_tipo') == "EJA":
        df = pd.read_excel(file_path, sheet_name=0, header=None, skiprows=1)
        df['RM_str'] = df.iloc[:, 2].apply(lambda x: str(int(x)) if pd.notna(x) and float(x) != 0 else "")
        df['NOME'] = df.iloc[:, 3]
        df['NASC.'] = df.iloc[:, 6]
        def get_ra(row):
            try:
                val = row.iloc[7]
                if pd.isna(val) or float(val) == 0:
                    return row.iloc[8]
                else:
                    return val
            except:
                return row.iloc[7]
        df['RA'] = df.apply(get_ra, axis=1)
        df['SÉRIE'] = df.iloc[:, 0]
        alunos = df[df['RM_str'] != ""][['RM_str', 'NOME']].drop_duplicates()
    else:
        planilha = pd.read_excel(file_path, sheet_name='LISTA CORRIDA')
        def format_rm(x):
            try:
                return str(int(float(x)))
            except:
                return str(x)
        planilha['RM_str'] = planilha['RM'].apply(format_rm)
        alunos = planilha[planilha['RM_str'] != "0"][['RM_str', 'NOME']].drop_duplicates()
    options_html = ""
    for _, row in alunos.iterrows():
        rm_str = row['RM_str']
        nome = row['NOME']
        options_html += f'<option value="{rm_str}">{rm_str} - {nome}</option>'
    if request.method == 'POST':
        rm = request.form.get('rm')
        tipo = request.form.get('tipo')
        if not rm or not tipo:
            flash("Selecione o aluno e o tipo de declaração.", "error")
            return redirect(url_for('declaracao_select'))
        declaracao_html = gerar_declaracao_escolar(file_path, rm, tipo)
        if declaracao_html is None:
            flash("Aluno não encontrado.", "error")
            return redirect(url_for('declaracao_select'))
        return declaracao_html
    select_form = f'''
    <!doctype html>
    <html lang="pt-br">
    <head>
      <meta charset="utf-8">
      <title>Declaração Escolar - Seleção de Aluno</title>
      <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
      <link href="https://cdnjs.cloudflare.com/ajax/libs/select2/4.0.13/css/select2.min.css" rel="stylesheet" />
      <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
      <style>
        body {{
          background: #eef2f3;
          font-family: 'Montserrat', sans-serif;
        }}
        header {{
          background: linear-gradient(90deg, #283E51, #4B79A1);
          color: #fff;
          padding: 20px;
          text-align: center;
          border-bottom: 3px solid #1d2d3a;
          border-radius: 0 0 15px 15px;
          box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .container-form {{
          background: #fff;
          padding: 40px;
          border-radius: 10px;
          box-shadow: 0 4px 12px rgba(0,0,0,0.15);
          margin: 40px auto;
          max-width: 600px;
        }}
        .btn-primary {{
          background-color: #283E51;
          border: none;
        }}
        .btn-primary:hover {{
          background-color: #1d2d3a;
        }}
        footer {{
          background-color: #424242;
          color: #fff;
          text-align: center;
          padding: 10px;
          position: fixed;
          bottom: 0;
          width: 100%;
        }}
      </style>
    </head>
    <body>
      <header>
        <h1>Declaração Escolar</h1>
        <p>Selecione o aluno e o tipo de declaração</p>
      </header>
      <div class="container container-form">
        <form method="POST" onsubmit="return confirmDeclaration();">
          <div class="form-group">
            <label for="rm">Aluno:</label>
            <select class="form-control" id="rm" name="rm" required>
              <option value="">Selecione</option>
              {options_html}
            </select>
          </div>
          <div class="form-group">
            <label for="tipo">Tipo de Declaração:</label>
            <select class="form-control" id="tipo" name="tipo" required>
              <option value="">Selecione</option>
              <option value="Escolaridade">Declaração de Escolaridade</option>
              <option value="Transferencia">Declaração de Transferência</option>
              <option value="Conclusão">Declaração de Conclusão</option>
            </select>
          </div>
          <button type="submit" class="btn btn-primary">Gerar Declaração</button>
        </form>
      </div>
      <footer>
        Desenvolvido por Nilson Cruz © 2025. Todos os direitos reservados.
      </footer>
      <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
      <script src="https://cdnjs.cloudflare.com/ajax/libs/select2/4.0.13/js/select2.min.js"></script>
      <script>
        $(document).ready(function() {{
          $('#rm').select2({{
            placeholder: "Selecione o aluno",
            allowClear: true
          }});
        }});
        function confirmDeclaration() {{
            var tipo = document.getElementById('tipo').value;
            if(tipo === "Transferencia") {{
                return confirm("Você está gerando uma declaração de transferência, essa é a declaração correta a ser gerada?");
            }}
            return true;
        }}
      </script>
    </body>
    </html>
    '''
    return render_template_string(select_form)


@app.route('/declaracao/tipo', methods=['GET', 'POST'])
@login_required
def declaracao_tipo():
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        if tipo == 'Fundamental':
            return redirect(url_for('declaracao_upload'))
        elif tipo == 'EJA':
            return redirect(url_for('declaracao_upload_eja'))
    form_html = '''
    <!doctype html>
    <html lang="pt-br">
    <head>
         <meta charset="utf-8">
         <title>Selecionar Tipo de Declaração Escolar</title>
         <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
         <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
         <style>
         body {
             background: #eef2f3;
             font-family: 'Montserrat', sans-serif;
         }
         header {
             background: linear-gradient(90deg, #283E51, #4B79A1);
             color: #fff;
             padding: 20px;
             text-align: center;
             border-bottom: 3px solid #1d2d3a;
         }
         .container-form {
             background: #fff;
             padding: 40px;
             border-radius: 10px;
             box-shadow: 0 4px 12px rgba(0,0,0,0.15);
             margin: 40px auto;
             max-width: 600px;
         }
         .btn-primary {
             background-color: #283E51;
             border: none;
         }
         .btn-primary:hover {
             background-color: #1d2d3a;
         }
         footer {
             background-color: #424242;
             color: #fff;
             text-align: center;
             padding: 10px;
             position: fixed;
             bottom: 0;
             width: 100%;
         }
         </style>
    </head>
    <body>
         <header>
             <h1>Selecionar Tipo de Declaração Escolar</h1>
         </header>
         <div class="container-form">
             <form method="POST">
                 <div class="form-group">
                     <label for="tipo">Selecione o tipo:</label>
                     <select class="form-control" id="tipo" name="tipo" required>
                         <option value="">Selecione</option>
                         <option value="Fundamental">Declaração Fundamental</option>
                         <option value="EJA">Declaração EJA</option>
                     </select>
                 </div>
                 <button type="submit" class="btn btn-primary">Continuar</button>
             </form>
         </div>
         <footer>
             Desenvolvido por Nilson Cruz © 2025. Todos os direitos reservados.
         </footer>
    </body>
    </html>
    '''
    return render_template_string(form_html)


@app.route('/upload_foto', methods=['POST'])
def upload_foto():
    if 'foto_file' not in request.files:
        return "Nenhum arquivo de foto enviado", 400
    rm = request.form.get('rm')
    if not rm:
        return "RM não fornecido", 400
    file = request.files['foto_file']
    if file.filename == '':
        return "Nenhuma foto selecionada", 400
    if not allowed_file(file.filename):
        return "Formato de imagem não permitido", 400
    original_filename = secure_filename(file.filename)
    _, ext = os.path.splitext(original_filename)
    new_filename = secure_filename(f"{rm}{ext.lower()}")
    file_path = os.path.join('static', 'fotos', new_filename)
    file.save(file_path)
    flash("Foto anexada com sucesso", "success")
    return redirect(url_for('carteirinhas'))


@app.route('/upload_multiplas_fotos', methods=['POST'])
def upload_multiplas_fotos():
    rms = request.form.getlist("rm[]")
    files = request.files.getlist("foto_file[]")
    if not files:
        return "Nenhuma foto enviada", 400
    for rm, file in zip(rms, files):
        if file.filename == '' or not rm or not allowed_file(file.filename):
            continue
        original_filename = secure_filename(file.filename)
        _, ext = os.path.splitext(original_filename)
        new_filename = secure_filename(f"{rm}{ext.lower()}")
        file_path = os.path.join('static', 'fotos', new_filename)
        file.save(file_path)
    flash("Foto(s) anexada(s) com sucesso", "success")
    return redirect(url_for('carteirinhas'))


@app.route('/upload_inline_foto', methods=['POST'])
def upload_inline_foto():
    if 'foto_file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    rm = request.form.get('rm')
    if not rm:
        return jsonify({'error': 'RM não fornecido'}), 400
    file = request.files['foto_file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'error': 'Formato de imagem não permitido'}), 400
    original_filename = secure_filename(file.filename)
    _, ext = os.path.splitext(original_filename)
    new_filename = secure_filename(f"{rm}{ext.lower()}")
    file_path = os.path.join('static', 'fotos', new_filename)
    file.save(file_path)
    return jsonify({'url': f"/static/fotos/{new_filename}", 'message': "Foto anexada com sucesso"})

@app.route('/quadros')
@login_required
def quadros():
    return render_template('quadros.html')

@app.route('/quadros/inclusao', methods=['GET', 'POST'])
@login_required
def quadros_inclusao():
    if request.method == 'POST':
        # Atualiza as listas na sessão (Fundamental e EJA)
        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')

        if fundamental_file and fundamental_file.filename != '':
            fundamental_filename = secure_filename(f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename)
            fundamental_path = os.path.join(app.config['UPLOAD_FOLDER'], fundamental_filename)
            fundamental_file.save(fundamental_path)
            session['lista_fundamental'] = fundamental_path

        if eja_file and eja_file.filename != '':
            eja_filename = secure_filename(f"eja_{uuid.uuid4().hex}_" + eja_file.filename)
            eja_path = os.path.join(app.config['UPLOAD_FOLDER'], eja_filename)
            eja_file.save(eja_path)
            session['lista_eja'] = eja_path

        # Carrega as listas piloto
        df_fundamental = None
        df_eja = None

        if session.get('lista_fundamental'):
            try:
                with open(session['lista_fundamental'], 'rb') as f_fund:
                    df_fundamental = pd.read_excel(f_fund, sheet_name="LISTA CORRIDA")
            except Exception:
                flash("Erro ao ler a Lista Piloto Fundamental.", "error")
                return redirect(url_for('quadros_inclusao'))

        if session.get('lista_eja'):
            try:
                with open(session['lista_eja'], 'rb') as f_eja:
                    df_eja = pd.read_excel(f_eja, sheet_name="LISTA CORRIDA")
            except Exception:
                flash("Erro ao ler a Lista Piloto EJA.", "error")
                return redirect(url_for('quadros_inclusao'))

        if df_fundamental is None and df_eja is None:
            flash("Nenhuma lista piloto disponível.", "error")
            return redirect(url_for('quadros_inclusao'))

        # Abre o modelo
        model_path = os.path.join("modelos", "Quadro de Alunos com Deficiência - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Inclusão não encontrado.", "error")
            return redirect(url_for('quadros_inclusao'))
        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo de inclusão: {str(e)}", "error")
            return redirect(url_for('quadros_inclusao'))

        ws = wb.active
        set_merged_cell_value(ws, "C3", "Luciana Rocha Augustinho")
        set_merged_cell_value(ws, "H3", "Ana Carolina Valencio da Silva Rodrigues")
        set_merged_cell_value(ws, "K3", "Rosemeire de Souza Pereira")
        set_merged_cell_value(ws, "C4", "Rafael Marques Lima")
        set_merged_cell_value(ws, "H4", "Rita de Cassia de Andrade")
        set_merged_cell_value(ws, "K4", "Ana Paula Rodrigues de Assis Santos")
        set_merged_cell_value(ws, "P4", datetime.now().strftime("%d/%m/%Y"))

        start_row = 7
        current_row = start_row

        # Processa alunos da Lista Piloto Fundamental
        if df_fundamental is not None:
            if len(df_fundamental.columns) < 21:
                flash("O arquivo da Lista Piloto Fundamental não possui colunas suficientes.", "error")
                return redirect(url_for('quadros_inclusao'))

            inclusion_col_fund = df_fundamental.columns[13]
            for idx, row in df_fundamental.iterrows():
                if not str(row['RM']).strip() or str(row['RM']).strip() == "0":
                    continue

                if str(row[inclusion_col_fund]).strip().lower() == "sim":
                    col_a_val = str(row[df_fundamental.columns[0]]).strip()
                    match = re.match(r"(\d+º).*?([A-Za-z])$", col_a_val)
                    if match:
                        nivel = match.group(1)
                        turma = match.group(2)
                    else:
                        nivel = col_a_val
                        turma = ""

                    horario = str(row[df_fundamental.columns[10]]).strip()
                    if "08h" in horario and "12h" in horario:
                        periodo = "MANHÃ"
                    elif horario == "13h30 às 17h30":
                        periodo = "TARDE"
                    elif horario == "19h00 às 23h00":
                        periodo = "NOITE"
                    else:
                        periodo = ""

                    nome_aluno = str(row[df_fundamental.columns[3]]).strip()
                    data_nasc = row[df_fundamental.columns[5]]
                    if pd.notna(data_nasc):
                        try:
                            data_nasc = pd.to_datetime(data_nasc, errors='coerce')
                            if pd.notna(data_nasc):
                                data_nasc = data_nasc.strftime('%d/%m/%Y')
                            else:
                                data_nasc = "Desconhecida"
                        except:
                            data_nasc = "Desconhecida"
                    else:
                        data_nasc = "Desconhecida"

                    professor = str(row[df_fundamental.columns[14]]).strip()
                    plano = str(row[df_fundamental.columns[15]]).strip()
                    aee = str(row[df_fundamental.columns[16]]).strip() if len(df_fundamental.columns) > 16 else ""
                    deficiencia = str(row[df_fundamental.columns[17]]).strip() if len(df_fundamental.columns) > 17 else ""
                    observacoes = str(row[df_fundamental.columns[18]]).strip() if len(df_fundamental.columns) > 18 else ""
                    cadeira = str(row[df_fundamental.columns[19]]).strip() if len(df_fundamental.columns) > 19 else ""
                    adequacoes = "-"
                    extra_field = row[df_fundamental.columns[20]] if len(df_fundamental.columns) > 20 else ""

                    ws.cell(row=current_row, column=2, value=nivel)
                    ws.cell(row=current_row, column=3, value=turma)
                    ws.cell(row=current_row, column=4, value=periodo)
                    ws.cell(row=current_row, column=5, value=horario)
                    ws.cell(row=current_row, column=6, value=nome_aluno)
                    ws.cell(row=current_row, column=7, value=data_nasc)
                    ws.cell(row=current_row, column=8, value=professor)
                    ws.cell(row=current_row, column=9, value=plano)
                    ws.cell(row=current_row, column=10, value=aee)
                    ws.cell(row=current_row, column=11, value=deficiencia)
                    ws.cell(row=current_row, column=12, value=observacoes)
                    ws.cell(row=current_row, column=13, value=cadeira)
                    ws.cell(row=current_row, column=14, value=adequacoes)
                    ws.cell(row=current_row, column=15, value=extra_field)
                    current_row += 1

        # Processa alunos da Lista Piloto EJA com novo mapeamento
        if df_eja is not None:
            if len(df_eja.columns) < 25:
                flash("O arquivo da Lista Piloto EJA não possui colunas suficientes.", "error")
                return redirect(url_for('quadros_inclusao'))

            inclusion_col_eja = df_eja.columns[17]
            for idx, row in df_eja.iterrows():
                if not str(row['RM']).strip() or str(row['RM']).strip() == "0":
                    continue

                if str(row[inclusion_col_eja]).strip().lower() == "sim":
                    nivel = str(row[df_eja.columns[0]]).strip()
                    turma = "A"
                    periodo = "NOITE"
                    horario = str(row[df_eja.columns[15]]).strip()
                    nome_aluno = str(row[df_eja.columns[3]]).strip()
                    data_nasc = row[df_eja.columns[6]]
                    if pd.notna(data_nasc):
                        try:
                            data_nasc = pd.to_datetime(data_nasc, errors='coerce')
                            if pd.notna(data_nasc):
                                data_nasc = data_nasc.strftime('%d/%m/%Y')
                            else:
                                data_nasc = "Desconhecida"
                        except:
                            data_nasc = "Desconhecida"
                    else:
                        data_nasc = "Desconhecida"
                    professor = str(row[df_eja.columns[18]]).strip()
                    plano = str(row[df_eja.columns[19]]).strip()
                    aee = str(row[df_eja.columns[20]]).strip() if len(df_eja.columns) > 20 else ""
                    deficiencia = str(row[df_eja.columns[21]]).strip() if len(df_eja.columns) > 21 else ""
                    observacoes = str(row[df_eja.columns[22]]).strip() if len(df_eja.columns) > 22 else ""
                    cadeira = str(row[df_eja.columns[23]]).strip() if len(df_eja.columns) > 23 else ""
                    adequacoes = "-"
                    atendimentos_saude = str(row[df_eja.columns[24]]).strip() if len(df_eja.columns) > 24 else ""
                    ws.cell(row=current_row, column=2, value=nivel)
                    ws.cell(row=current_row, column=3, value=turma)
                    ws.cell(row=current_row, column=4, value=periodo)
                    ws.cell(row=current_row, column=5, value=horario)
                    ws.cell(row=current_row, column=6, value=nome_aluno)
                    ws.cell(row=current_row, column=7, value=data_nasc)
                    ws.cell(row=current_row, column=8, value=professor)
                    ws.cell(row=current_row, column=9, value=plano)
                    ws.cell(row=current_row, column=10, value=aee)
                    ws.cell(row=current_row, column=11, value=deficiencia)
                    ws.cell(row=current_row, column=12, value=observacoes)
                    ws.cell(row=current_row, column=13, value=cadeira)
                    ws.cell(row=current_row, column=14, value=adequacoes)
                    ws.cell(row=current_row, column=15, value=atendimentos_saude)
                    current_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        meses = {1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril", 5: "maio", 6: "junho",
                 7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
        mes = meses[datetime.now().month].capitalize()
        filename = f"Quadro de Inclusão - {mes} - E.M José Padin Mouta.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        # Para GET, renderize o template separado
        return render_template("quadros_inclusao.html")

@app.route('/quadros/atendimento_mensal', methods=['GET', 'POST'])
@login_required
def quadro_atendimento_mensal():
    if request.method == 'POST':
        # Recebe os arquivos enviados pelo formulário
        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')

        if fundamental_file and fundamental_file.filename != '':
            filename = secure_filename(fundamental_file.filename)
            unique_filename = f"atendimento_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            fundamental_file.save(file_path)
            session['lista_fundamental'] = file_path

        if eja_file and eja_file.filename != '':
            filename = secure_filename(eja_file.filename)
            unique_filename = f"atendimento_eja_{uuid.uuid4().hex}_{filename}"
            file_path_eja = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            eja_file.save(file_path_eja)
            session['lista_eja'] = file_path_eja

        # Abre o arquivo da lista fundamental
        file_path = session.get('lista_fundamental')
        if file_path and os.path.exists(file_path):
            lista_file = open(file_path, 'rb')
        else:
            lista_file = None

        if not lista_file:
            flash("Nenhum arquivo enviado.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        # Verifica a existência do modelo Excel
        model_path = os.path.join("modelos", "Quadro de Atendimento Mensal - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo Atendimento Mensal não encontrado.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        try:
            with open(model_path, "rb") as f:
                wb_modelo = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo de atendimento mensal: {str(e)}", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        # Seleciona a planilha do modelo (segunda se houver mais de uma)
        if len(wb_modelo.worksheets) > 1:
            ws_modelo = wb_modelo.worksheets[1]
        else:
            ws_modelo = wb_modelo.active

        # Preenche dados fixos no modelo
        set_merged_cell_value(ws_modelo, "B5", "E.M José Padin Mouta")
        set_merged_cell_value(ws_modelo, "C6", "Rafael Fernando da Silva")
        set_merged_cell_value(ws_modelo, "B7", "46034")
        current_month = datetime.now().strftime("%m")
        set_merged_cell_value(ws_modelo, "A13", f"{current_month}/2025")

        try:
            lista_file.seek(0)
            wb_lista = load_workbook(lista_file, data_only=True)
        except Exception:
            flash("Erro ao ler o arquivo da lista piloto.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        # Procura a aba "Total de Alunos" na lista
        sheet_name = None
        for name in wb_lista.sheetnames:
            if name.strip().lower() == "total de alunos":
                sheet_name = name
                break
        if not sheet_name:
            flash("A aba 'Total de Alunos' não foi encontrada na lista piloto.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        ws_total = wb_lista[sheet_name]

        # Preenche blocos do modelo com dados da lista FUNDAMENTAL
        for r, source_row in zip(range(55, 57), range(13, 15)):
            value_B = ws_total.cell(row=source_row, column=7).value
            value_C = ws_total.cell(row=source_row, column=8).value
            set_merged_cell_value(ws_modelo, f"B{r}", value_B)
            set_merged_cell_value(ws_modelo, f"C{r}", value_C)
            set_merged_cell_value(ws_modelo, f"D{r}", f"=B{r}+C{r}")

        for r, source_row in zip(range(57, 61), range(15, 19)):
            value_B = ws_total.cell(row=source_row, column=7).value
            value_C = ws_total.cell(row=source_row, column=8).value
            set_merged_cell_value(ws_modelo, f"B{r}", value_B)
            set_merged_cell_value(ws_modelo, f"C{r}", value_C)
            set_merged_cell_value(ws_modelo, f"D{r}", f"=B{r}+C{r}")

        for r, source_row in zip(range(73, 80), range(20, 27)):
            value_B = ws_total.cell(row=source_row, column=7).value
            value_C = ws_total.cell(row=source_row, column=8).value
            set_merged_cell_value(ws_modelo, f"B{r}", value_B)
            set_merged_cell_value(ws_modelo, f"C{r}", value_C)
            set_merged_cell_value(ws_modelo, f"D{r}", f"=B{r}+C{r}")

        for r, source_row in zip(range(91, 98), range(28, 35)):
            value_B = ws_total.cell(row=source_row, column=7).value
            value_C = ws_total.cell(row=source_row, column=8).value
            set_merged_cell_value(ws_modelo, f"B{r}", value_B)
            set_merged_cell_value(ws_modelo, f"C{r}", value_C)
            set_merged_cell_value(ws_modelo, f"D{r}", f"=B{r}+C{r}")

        # Preenchimento de campos específicos
        value_R20 = ws_total.cell(row=37, column=9).value
        set_merged_cell_value(ws_modelo, "R20", value_R20)
        set_merged_cell_value(ws_modelo, "R24", "-")
        value_R28 = ws_total.cell(row=39, column=9).value
        set_merged_cell_value(ws_modelo, "R28", value_R28)

        set_merged_cell_value(ws_modelo, "B37", ws_total.cell(row=6, column=7).value)
        set_merged_cell_value(ws_modelo, "B38", ws_total.cell(row=7, column=7).value)
        set_merged_cell_value(ws_modelo, "B39", ws_total.cell(row=8, column=7).value)
        set_merged_cell_value(ws_modelo, "B40", ws_total.cell(row=9, column=7).value)
        set_merged_cell_value(ws_modelo, "B41", ws_total.cell(row=10, column=7).value)
        set_merged_cell_value(ws_modelo, "B42", ws_total.cell(row=11, column=7).value)

        set_merged_cell_value(ws_modelo, "C37", ws_total.cell(row=6, column=8).value)
        set_merged_cell_value(ws_modelo, "C38", ws_total.cell(row=7, column=8).value)
        set_merged_cell_value(ws_modelo, "C39", ws_total.cell(row=8, column=8).value)
        set_merged_cell_value(ws_modelo, "C40", ws_total.cell(row=9, column=8).value)
        set_merged_cell_value(ws_modelo, "C41", ws_total.cell(row=10, column=8).value)
        set_merged_cell_value(ws_modelo, "C42", ws_total.cell(row=11, column=8).value)

        # Processa dados da Lista Piloto EJA
        eja_path = session.get('lista_eja')
        if not eja_path or not os.path.exists(eja_path):
            flash("Arquivo da Lista Piloto EJA não encontrado.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))
        with open(eja_path, 'rb') as f_eja:
            wb_eja = load_workbook(f_eja, data_only=True)
        sheet_name_eja = None
        for name in wb_eja.sheetnames:
            if name.strip().lower() == "total de alunos":
                sheet_name_eja = name
                break
        if not sheet_name_eja:
            flash("A aba 'Total de Alunos' não foi encontrada na Lista Piloto EJA.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))
        ws_total_eja = wb_eja[sheet_name_eja]

        set_merged_cell_value(ws_modelo, "L19", ws_total_eja.cell(row=6, column=5).value)
        set_merged_cell_value(ws_modelo, "L20", ws_total_eja.cell(row=7, column=5).value)
        set_merged_cell_value(ws_modelo, "L21", ws_total_eja.cell(row=8, column=5).value)
        set_merged_cell_value(ws_modelo, "L22", ws_total_eja.cell(row=9, column=5).value)
        set_merged_cell_value(ws_modelo, "M19", ws_total_eja.cell(row=6, column=6).value)
        set_merged_cell_value(ws_modelo, "M20", ws_total_eja.cell(row=7, column=6).value)
        set_merged_cell_value(ws_modelo, "M21", ws_total_eja.cell(row=8, column=6).value)
        set_merged_cell_value(ws_modelo, "M22", ws_total_eja.cell(row=9, column=6).value)
        set_merged_cell_value(ws_modelo, "L27", ws_total_eja.cell(row=11, column=5).value)
        set_merged_cell_value(ws_modelo, "L28", ws_total_eja.cell(row=12, column=5).value)
        set_merged_cell_value(ws_modelo, "L29", ws_total_eja.cell(row=13, column=5).value)
        set_merged_cell_value(ws_modelo, "L30", ws_total_eja.cell(row=14, column=5).value)
        set_merged_cell_value(ws_modelo, "M27", ws_total_eja.cell(row=11, column=6).value)
        set_merged_cell_value(ws_modelo, "M28", ws_total_eja.cell(row=12, column=6).value)
        set_merged_cell_value(ws_modelo, "M29", ws_total_eja.cell(row=13, column=6).value)
        set_merged_cell_value(ws_modelo, "M30", ws_total_eja.cell(row=14, column=6).value)
        set_merged_cell_value(ws_modelo, "L35", ws_total_eja.cell(row=16, column=5).value)
        set_merged_cell_value(ws_modelo, "L36", ws_total_eja.cell(row=17, column=5).value)
        set_merged_cell_value(ws_modelo, "L37", ws_total_eja.cell(row=18, column=5).value)
        set_merged_cell_value(ws_modelo, "M35", ws_total_eja.cell(row=16, column=6).value)
        set_merged_cell_value(ws_modelo, "M36", ws_total_eja.cell(row=17, column=6).value)
        set_merged_cell_value(ws_modelo, "M37", ws_total_eja.cell(row=18, column=6).value)
        set_merged_cell_value(ws_modelo, "R32", ws_total_eja.cell(row=20, column=7).value)
        set_merged_cell_value(ws_modelo, "R24", "-")

        output = BytesIO()
        wb_modelo.save(output)
        output.seek(0)
        filename = f"Quadro de Atendimento Mensal - {datetime.now().strftime('%d%m')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        # Para GET, renderiza o template separado
        return render_template("quadro_atendimento_mensal.html")

@app.route('/quadros/transferencias', methods=['GET', 'POST'])
@login_required
def quadro_transferencias():
    if request.method == 'POST':
        # Obtém dados do formulário
        period_start_str = request.form.get('period_start')
        period_end_str = request.form.get('period_end')
        responsavel = request.form.get('responsavel')

        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')

        if not period_start_str or not period_end_str or not responsavel:
            flash("Por favor, preencha todos os campos.", "error")
            return redirect(url_for('quadro_transferencias'))

        # Salva/atualiza a Lista Piloto FUNDAMENTAL
        if fundamental_file and fundamental_file.filename != '':
            fundamental_filename = secure_filename(f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename)
            fundamental_path = os.path.join(app.config['UPLOAD_FOLDER'], fundamental_filename)
            fundamental_file.save(fundamental_path)
            session['lista_fundamental'] = fundamental_path
        else:
            fundamental_path = session.get('lista_fundamental')
            if not fundamental_path or not os.path.exists(fundamental_path):
                flash("Lista Piloto Fundamental não encontrada.", "error")
                return redirect(url_for('quadro_transferencias'))

        # Salva/atualiza a Lista Piloto EJA (opcional)
        if eja_file and eja_file.filename != '':
            eja_filename = secure_filename(f"eja_{uuid.uuid4().hex}_" + eja_file.filename)
            eja_path = os.path.join(app.config['UPLOAD_FOLDER'], eja_filename)
            eja_file.save(eja_path)
            session['lista_eja'] = eja_path
        else:
            eja_path = session.get('lista_eja')
            # Se não houver, prossegue sem erro

        try:
            period_start = datetime.strptime(period_start_str, "%Y-%m-%d")
            period_end = datetime.strptime(period_end_str, "%Y-%m-%d")
        except Exception:
            flash("Formato de data inválido.", "error")
            return redirect(url_for('quadro_transferencias'))

        # ---- PARTE 1: Processa a Lista Piloto FUNDAMENTAL
        try:
            df_fundamental = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
        except Exception as e:
            flash(f"Erro ao ler a Lista Piloto Fundamental: {str(e)}", "error")
            return redirect(url_for('quadro_transferencias'))

        motivo_map = {
            "Dentro da Rede": "Dentro da rede",
            "Rede Estadual": "Dentro da rede",
            "Litoral": "Mudança de Municipio",
            "Mudança de Municipio": "Mudança de Municipio",
            "São Paulo": "Mudança de Municipio",
            "ABCD": "Mudança de Municipio",
            "Interior": "Mudança de Municipio",
            "Outros Estados": "Mudança de estado",
            "Particular": "Mudança para Escola Particular",
            "País": "Mudança de País"
        }

        transfer_records = []
        col_V_index = 21  # índice 0-based

        for idx, row in df_fundamental.iterrows():
            if len(row) < 9:
                continue

            obs_value = str(row.iloc[8]) if len(row) > 8 else ""
            motivo_raw = str(row.iloc[col_V_index]).strip() if len(row) > col_V_index else ""
            motivo_w = str(row.iloc[22]).strip() if len(row) > 22 else ""

            match = re.search(r"(TE)\s*(\d{1,2}/\d{1,2})", obs_value)
            if match:
                te_date_str = match.group(2)
                te_date_full_str = f"{te_date_str}/{period_start.year}"
                try:
                    te_date = datetime.strptime(te_date_full_str, "%d/%m/%Y")
                except:
                    continue
                if period_start <= te_date <= period_end:
                    nome = str(row.iloc[3])
                    dn_val = row.iloc[5]
                    dn_str = ""
                    if pd.notna(dn_val):
                        try:
                            dn_dt = pd.to_datetime(dn_val, errors='coerce')
                            if pd.notna(dn_dt):
                                dn_str = dn_dt.strftime('%d/%m/%y')
                        except:
                            dn_str = ""
                    ra = str(row.iloc[6])
                    situacao = "Parcial"
                    breda = "Não"
                    nivel_classe = str(row.iloc[0])
                    tipo_field = "TE"
                    if motivo_raw in motivo_map:
                        reason_final = motivo_map[motivo_raw]
                    else:
                        reason_final = motivo_raw
                    if motivo_w:
                        reason_final = f"{reason_final} ({motivo_w})"
                    remanejamento = "-"
                    data_te = te_date.strftime("%d/%m/%Y")
                    record = {
                        "nome": nome,
                        "dn": dn_str,
                        "ra": ra,
                        "situacao": situacao,
                        "breda": breda,
                        "nivel_classe": nivel_classe,
                        "tipo": tipo_field,
                        "observacao": reason_final,
                        "remanejamento": remanejamento,
                        "data": data_te
                    }
                    transfer_records.append(record)

        # ---- PARTE 2: Processa a Lista Piloto EJA
        if eja_path and os.path.exists(eja_path):
            try:
                df_eja = pd.read_excel(eja_path, sheet_name="LISTA CORRIDA")
            except Exception as e:
                flash(f"Erro ao ler a Lista Piloto EJA: {str(e)}", "error")
                return redirect(url_for('quadro_transferencias'))

            for idx, row in df_eja.iterrows():
                if len(row) < 11:
                    continue
                col_k_value = str(row.iloc[10]).strip() if len(row) > 10 else ""
                if not col_k_value:
                    continue
                match_eja = re.search(r"(TE|MC|MCC)\s*(\d{1,2}/\d{1,2})", col_k_value, re.IGNORECASE)
                if match_eja:
                    tipo_str = match_eja.group(1).upper()
                    date_str = match_eja.group(2)
                    eja_date_full = f"{date_str}/{period_start.year}"
                    try:
                        eja_date = datetime.strptime(eja_date_full, "%d/%m/%Y")
                    except:
                        continue
                    if period_start <= eja_date <= period_end:
                        nome = str(row.iloc[3])
                        dn_val = row.iloc[6]
                        dn_str = ""
                        if pd.notna(dn_val):
                            try:
                                dn_dt = pd.to_datetime(dn_val, errors='coerce')
                                if pd.notna(dn_dt):
                                    dn_str = dn_dt.strftime('%d/%m/%Y')
                            except:
                                dn_str = ""
                        ra_val = row.iloc[7]
                        if pd.isna(ra_val) or (isinstance(ra_val, (int, float)) and float(ra_val) == 0):
                            ra_val = row.iloc[8]
                        situacao = "Parcial"
                        breda = "Não"
                        nivel_classe = str(row.iloc[0])
                        tipo_field = tipo_str
                        if tipo_field in ["MC", "MCC"]:
                            obs_final = "Desistencia"
                        else:
                            part_z = str(row.iloc[25]).strip() if len(row) > 25 else ""
                            part_aa = str(row.iloc[26]).strip() if len(row) > 26 else ""
                            if part_aa:
                                obs_final = f"{part_z} ({part_aa})".strip()
                            else:
                                obs_final = part_z
                        remanejamento = "-"
                        data_te = eja_date.strftime("%d/%m/%Y")
                        record = {
                            "nome": nome,
                            "dn": dn_str,
                            "ra": str(ra_val),
                            "situacao": situacao,
                            "breda": breda,
                            "nivel_classe": nivel_classe,
                            "tipo": tipo_field,
                            "observacao": obs_final,
                            "remanejamento": remanejamento,
                            "data": data_te
                        }
                        transfer_records.append(record)

        if not transfer_records:
            flash("Nenhum registro de TE/MC/MCC encontrado no período especificado.", "error")
            return redirect(url_for('quadro_transferencias'))

        model_path = os.path.join("modelos", "Quadro Informativo - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Quadro Informativo (Transferências) não encontrado.", "error")
            return redirect(url_for('quadro_transferencias'))

        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo: {str(e)}", "error")
            return redirect(url_for('quadro_transferencias'))

        ws = wb.active
        set_merged_cell_value(ws, "B9", responsavel)
        set_merged_cell_value(ws, "J9", datetime.now().strftime("%d/%m/%Y"))

        start_row = 12
        current_row = start_row

        for record in transfer_records:
            set_merged_cell_value(ws, f"A{current_row}", record["nome"])
            set_merged_cell_value(ws, f"B{current_row}", record["dn"])
            set_merged_cell_value(ws, f"C{current_row}", record["ra"])
            set_merged_cell_value(ws, f"D{current_row}", record["situacao"])
            set_merged_cell_value(ws, f"E{current_row}", record["breda"])
            set_merged_cell_value(ws, f"F{current_row}", record["nivel_classe"])
            set_merged_cell_value(ws, f"G{current_row}", record["tipo"])
            set_merged_cell_value(ws, f"H{current_row}", record["observacao"])
            set_merged_cell_value(ws, f"I{current_row}", record["remanejamento"])
            set_merged_cell_value(ws, f"J{current_row}", record["data"])
            current_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Quadro_de_Transferencias_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        # Método GET: renderiza o template
        return render_template("quadro_transferencias.html")

if __name__ == '__main__':
    app.run(debug=True)
